#!/usr/bin/env python3
"""
Batch pipeline:
1) Tile-expand image to target cm (default 100x100)
2) Free crop by origin + size
3) Add white border then black border
4) Export PNG / JPEG / TIFF（依輸出副檔名）

Merged from kuotu expand/crop + continue/edge add_borders.
"""

from __future__ import annotations

import argparse
import math
import re
import sys
from pathlib import Path

from PIL import Image, ImageOps

SUPPORTED_EXTENSIONS = {
    ".jpg",
    ".jpeg",
    ".png",
    ".tif",
    ".tiff",
    ".webp",
    ".bmp",
}

CM_PER_INCH = 2.54
DEFAULT_DPI = 300.0
DEFAULT_TARGET_W_CM = 100.0
DEFAULT_TARGET_H_CM = 100.0
DEFAULT_WHITE_CM = 0.17
DEFAULT_BLACK_CM = 0.08

# 尺碼數字對：45x55 / 45×55 / 45*55 / 45 X 55
_SIZE_PAIR = r"(\d+(?:\.\d+)?)\s*[xX×*]\s*(\d+(?:\.\d+)?)"

# 依優先序匹配；有單位時先取 cm，再退回純數字（視為 cm）或英吋換算
# 每項：(名稱, 正則, 單位) — 正則須含兩個捕獲組 (寬, 高)
SIZE_PATTERNS: tuple[tuple[str, re.Pattern[str], str], ...] = (
    # 括號內厘米：15.75 X 19.69英寸(40x50厘米)
    (
        "paren_cm",
        re.compile(rf"[\(（]\s*{_SIZE_PAIR}\s*(?:cm|CM|厘米)?\s*[\)）]"),
        "cm",
    ),
    # 雙側帶 cm：45cm*55cm / 40 cm x 50 cm
    (
        "cm_both",
        re.compile(
            r"(\d+(?:\.\d+)?)\s*(?:cm|CM|厘米)\s*[xX×*]\s*"
            r"(\d+(?:\.\d+)?)\s*(?:cm|CM|厘米)"
        ),
        "cm",
    ),
    # 後綴厘米：40x50厘米 / 40×50cm
    (
        "cm_suffix",
        re.compile(rf"{_SIZE_PAIR}\s*(?:cm|CM|厘米)"),
        "cm",
    ),
    # 生产属性標籤：尺码:45x55
    (
        "labeled",
        re.compile(rf"尺码\s*[:：]\s*{_SIZE_PAIR}"),
        "cm",
    ),
    # 英吋：18 x 22 inch / 39.37*59.08英寸
    (
        "inch",
        re.compile(rf"{_SIZE_PAIR}\s*(?:inch(?:es)?|in|英寸)"),
        "inch",
    ),
    # 純尺碼欄：45x55（視為 cm）
    (
        "plain",
        re.compile(rf"^{_SIZE_PAIR}\s*$"),
        "cm",
    ),
)

# 相容舊呼叫（生产属性內嵌尺码）
SIZE_RE = SIZE_PATTERNS[3][1]

SKU_HEADER_NAMES = frozenset(
    {"平台sku", "平台sku id", "平台skuid", "sku", "sku id", "skuid"}
)
SIZE_HEADER_NAMES = frozenset({"尺码", "尺碼", "size"})
ATTR_HEADER_NAMES = frozenset({"生产属性", "生產屬性"})

# 資料夾名 → 平台 SKU ID（依序匹配，取第一個命中）
# 例：SKU45189918413 / sku-45189918413 / 45189918413 / 11997725258-1
FOLDER_SKU_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"(?i)^SKU[\s_\-]*(\d+)$"),  # SKU4518… / SKU_4518… / SKU-4518…
    re.compile(r"(?i)^SKU[\s_\-]*(\d+)"),  # SKU4518…_xxx（前綴帶數字）
    re.compile(r"^(\d+)$"),  # 純數字資料夾
    re.compile(r"^(\d+)[-_].+"),  # 11997725258-1（-/_ 前為 SKU）
)


def cm_to_px(cm: float, dpi: float) -> int:
    if cm <= 0:
        return 0
    return max(1, int(round(cm / CM_PER_INCH * dpi)))


def get_dpi(img: Image.Image, fallback: float) -> tuple[float, bool]:
    """Return (dpi, used_fallback). Prefer image metadata; else fallback."""
    dpi_info = img.info.get("dpi")
    if dpi_info:
        if isinstance(dpi_info, (tuple, list)) and len(dpi_info) >= 1:
            value = float(dpi_info[0])
            if value > 0:
                return value, False
        elif isinstance(dpi_info, (int, float)) and float(dpi_info) > 0:
            return float(dpi_info), False

    try:
        exif = img.getexif()
        if exif:
            x_res = exif.get(282)
            y_res = exif.get(283)
            unit = exif.get(296, 2)
            res = x_res or y_res
            if res is not None:
                if hasattr(res, "numerator") and hasattr(res, "denominator"):
                    value = float(res.numerator) / float(res.denominator or 1)
                else:
                    value = float(res)
                if value > 0:
                    if unit == 3:
                        value = value * CM_PER_INCH
                    return value, False
    except Exception:
        pass

    return fallback, True


def prepare_for_expand(img: Image.Image) -> Image.Image:
    if img.mode == "P":
        if "transparency" in img.info:
            return img.convert("RGBA")
        return img.convert("RGB")
    if img.mode == "LA":
        return img.convert("RGBA")
    if img.mode == "1":
        return img.convert("RGB")
    if img.mode not in ("RGB", "RGBA", "L", "CMYK"):
        return img.convert("RGB")
    return img


def build_tiled_canvas(src: Image.Image, target_w: int, target_h: int) -> Image.Image:
    mode = src.mode if src.mode in ("RGB", "RGBA", "L") else "RGB"
    if src.mode != mode:
        src = src.convert(mode)
    canvas = Image.new(mode, (target_w, target_h))
    src_w, src_h = src.size
    if src_w <= 0 or src_h <= 0:
        raise ValueError("來源圖片尺寸無效")
    repeat_x = math.ceil(target_w / src_w)
    repeat_y = math.ceil(target_h / src_h)
    for y in range(repeat_y):
        for x in range(repeat_x):
            canvas.paste(src, (x * src_w, y * src_h))
    return canvas


def add_double_border(
    img: Image.Image,
    white_cm: float,
    black_cm: float,
    dpi: float,
) -> Image.Image:
    prepared = prepare_for_expand(img)
    white_px = cm_to_px(white_cm, dpi)
    black_px = cm_to_px(black_cm, dpi)
    if white_px:
        prepared = ImageOps.expand(prepared, border=white_px, fill="white")
    if black_px:
        prepared = ImageOps.expand(prepared, border=black_px, fill="black")
    return prepared


def collect_images(input_path: Path, recursive: bool) -> list[Path]:
    if input_path.is_file():
        if input_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise ValueError(f"不支援的格式: {input_path.suffix}")
        return [input_path]

    if not input_path.is_dir():
        raise ValueError(f"路徑不存在: {input_path}")

    pattern = "**/*" if recursive else "*"
    return [
        p
        for p in sorted(input_path.glob(pattern))
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
    ]


def default_output_dir(input_path: Path) -> Path:
    if input_path.is_file():
        return input_path.parent / "pipeline_out"
    return input_path.parent / f"{input_path.name}_out"


def extract_sku_from_folder_name(folder_name: str) -> str | None:
    """
    從資料夾名稱抽出平台 SKU ID。

    支援：SKU45189918413、sku-45189918413、45189918413、11997725258-1 等。
    「數字-後綴」時取連字號／底線前的數字為 SKU。
    """
    name = folder_name.strip()
    if not name:
        return None
    for pattern in FOLDER_SKU_PATTERNS:
        match = pattern.search(name)
        if match:
            return match.group(1)
    return None


def normalize_sku(value: object) -> str:
    """Normalize Excel SKU cell to a digit string."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    # Excel 也可能寫成 SKU4518…
    extracted = extract_sku_from_folder_name(text)
    return extracted if extracted else text


def parse_size_cm(attr: object) -> tuple[float, float] | None:
    """
    從尺碼文字抽出 (寬_cm, 高_cm)。

    支援純尺碼欄（45x55）、生产属性（尺码:45x55）、
    cm／英吋單位，以及括號內厘米（優先）。
    """
    if attr is None:
        return None
    text = str(attr).strip()
    if not text:
        return None
    for _name, pattern, unit in SIZE_PATTERNS:
        match = pattern.search(text)
        if not match:
            continue
        # 取前兩個非空捕獲組
        nums = [g for g in match.groups() if g is not None]
        if len(nums) < 2:
            continue
        w, h = float(nums[0]), float(nums[1])
        if unit == "inch":
            w *= CM_PER_INCH
            h *= CM_PER_INCH
        if w <= 0 or h <= 0:
            return None
        return w, h
    return None


def _norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def _find_col(headers: list[object], names: frozenset[str]) -> int | None:
    for i, cell in enumerate(headers):
        if _norm_header(cell) in names:
            return i
    return None


class SkuCatalog:
    __slots__ = ("sizes", "all_skus")

    def __init__(
        self,
        sizes: dict[str, tuple[float, float]],
        all_skus: set[str],
    ) -> None:
        self.sizes = sizes
        self.all_skus = all_skus


def load_sku_catalog(excel_path: Path) -> SkuCatalog:
    """
    讀取 Excel：第 1 欄（或表頭「平台SKU」）為 SKU；
    優先讀「尺码」欄（截圖為 D 欄；若無「是否匹配」則常為第 3 欄），
    否則回退解析「生产属性」。
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "缺少 openpyxl，請執行: pip install openpyxl"
        ) from exc

    excel_path = excel_path.resolve()
    if not excel_path.is_file():
        raise ValueError(f"Excel 不存在: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_col=16, values_only=True))
        if not rows:
            return SkuCatalog(sizes={}, all_skus=set())

        header = list(rows[0])
        sku_col = _find_col(header, SKU_HEADER_NAMES)
        size_col = _find_col(header, SIZE_HEADER_NAMES)
        attr_col = _find_col(header, ATTR_HEADER_NAMES)
        has_header = sku_col is not None or size_col is not None

        # 無表頭時：A=SKU；尺碼優先第 3 欄(C)，再試第 4 欄(D)；生产属性=B
        if sku_col is None:
            sku_col = 0
        if size_col is None and not has_header:
            # 使用者所述「第三列才是尺碼」；亦相容含「是否匹配」時的 D 欄
            size_col = 2
        if attr_col is None and not has_header:
            attr_col = 1

        data_rows = rows[1:] if has_header else rows
        # 舊表僅兩欄且表頭為「平台SKU ID」時仍跳過首列
        if not has_header and rows:
            first_sku = _norm_header(rows[0][0] if rows[0] else "")
            if first_sku in SKU_HEADER_NAMES or first_sku in {
                "平台skuid",
                "平台sku id",
            }:
                data_rows = rows[1:]

        sizes: dict[str, tuple[float, float]] = {}
        all_skus: set[str] = set()

        for row in data_rows:
            if not row:
                continue
            sku_raw = row[sku_col] if sku_col < len(row) else None
            sku = normalize_sku(sku_raw)
            if not sku:
                continue
            all_skus.add(sku)

            size: tuple[float, float] | None = None
            # 1) 專用尺碼欄
            if size_col is not None and size_col < len(row):
                size = parse_size_cm(row[size_col])
            # 2) 含「是否匹配」時若 C 欄不是尺碼，再試 D 欄
            if size is None and size_col == 2 and len(row) > 3:
                size = parse_size_cm(row[3])
            # 3) 回退生产属性
            if size is None and attr_col is not None and attr_col < len(row):
                size = parse_size_cm(row[attr_col])

            if size is not None:
                sizes[sku] = size

        return SkuCatalog(sizes=sizes, all_skus=all_skus)
    finally:
        wb.close()


def load_sku_sizes(excel_path: Path) -> dict[str, tuple[float, float]]:
    """
    Read Excel mapping: 平台SKU ID -> (crop_w_cm, crop_h_cm).

    Rows without a parseable 尺码 are omitted (caller should skip those folders).
    """
    return load_sku_catalog(excel_path).sizes


def build_dest(src: Path, input_path: Path, output_dir: Path) -> Path:
    if input_path.is_file():
        return output_dir / f"{src.stem}.tif"
    try:
        rel = src.relative_to(input_path)
    except ValueError:
        rel = Path(src.name)
    return (output_dir / rel).with_suffix(".tif")


def _under(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def save_expanded_image(image: Image.Image, dest: Path, dpi: float) -> None:
    """依副檔名輸出 PNG / JPEG / TIFF；未指定或未知副檔名時預設 TIFF。"""
    suffix = dest.suffix.lower()
    dest.parent.mkdir(parents=True, exist_ok=True)
    if suffix in {".jpg", ".jpeg"}:
        image.convert("RGB").save(dest, format="JPEG", quality=95, dpi=(dpi, dpi))
        return
    if suffix == ".png":
        image.save(dest, format="PNG", dpi=(dpi, dpi))
        return
    if suffix in {".tif", ".tiff", ""}:
        if not suffix:
            dest = dest.with_suffix(".tif")
        image.save(dest, format="TIFF", dpi=(dpi, dpi), compression="tiff_lzw")
        return
    raise ValueError(f"不支援的輸出格式：{suffix}（請用 .png / .jpg / .tif）")


def process_one(
    src: Path,
    dest: Path,
    *,
    fallback_dpi: float,
    target_w_cm: float,
    target_h_cm: float,
    crop_x: int,
    crop_y: int,
    crop_w_cm: float,
    crop_h_cm: float,
    white_cm: float,
    black_cm: float,
    use_image_dpi: bool = True,
    log=print,
) -> None:
    with Image.open(src) as img:
        n_frames = getattr(img, "n_frames", 1)
        if n_frames > 1:
            log(f"  提示: {src.name} 有 {n_frames} 頁，僅處理第一頁")
            img.seek(0)

        if use_image_dpi:
            dpi, used_fallback = get_dpi(img, fallback_dpi)
            if used_fallback:
                log(f"  提示: {src.name} 無 DPI 資訊，使用預設 {fallback_dpi:g} DPI")
        else:
            dpi = fallback_dpi

        img.load()
        prepared = prepare_for_expand(img)
        if prepared.mode not in ("RGB", "RGBA", "L"):
            prepared = prepared.convert("RGB")

        target_w = cm_to_px(target_w_cm, dpi)
        target_h = cm_to_px(target_h_cm, dpi)
        crop_w = cm_to_px(crop_w_cm, dpi)
        crop_h = cm_to_px(crop_h_cm, dpi)

        expanded = build_tiled_canvas(prepared, target_w, target_h)

        left = crop_x
        top = crop_y
        right = left + crop_w
        bottom = top + crop_h
        if left < 0 or top < 0:
            raise ValueError("裁切起點不可小於 0")
        if right > expanded.width or bottom > expanded.height:
            raise ValueError(
                f"裁切超出畫布。畫布={expanded.size}, "
                f"裁切框=({left}, {top}, {right}, {bottom})"
            )

        cropped = expanded.crop((left, top, right, bottom))
        result = add_double_border(cropped, white_cm=white_cm, black_cm=black_cm, dpi=dpi)

        if result.mode not in ("RGB", "L"):
            result = result.convert("RGB")

        save_expanded_image(result, dest, dpi)

        white_px = cm_to_px(white_cm, dpi)
        black_px = cm_to_px(black_cm, dpi)
        log(
            f"  OK  {src.name} -> {dest.name}  "
            f"(DPI={dpi:g}, 畫布={expanded.size}, 裁切={cropped.size}, "
            f"白={white_px}px, 黑={black_px}px)"
        )


def run_batch(
    input_path: Path,
    output_dir: Path | None = None,
    *,
    dpi: float = DEFAULT_DPI,
    target_w_cm: float = DEFAULT_TARGET_W_CM,
    target_h_cm: float = DEFAULT_TARGET_H_CM,
    crop_x: int = 0,
    crop_y: int = 0,
    crop_w_cm: float = 30.0,
    crop_h_cm: float = 30.0,
    white_cm: float = DEFAULT_WHITE_CM,
    black_cm: float = DEFAULT_BLACK_CM,
    recursive: bool = False,
    use_image_dpi: bool = True,
    log=print,
    on_progress=None,
) -> tuple[int, int]:
    input_path = input_path.resolve()
    out = (output_dir or default_output_dir(input_path)).resolve()

    if dpi <= 0:
        raise ValueError("DPI 必須大於 0")
    if target_w_cm <= 0 or target_h_cm <= 0:
        raise ValueError("擴展尺寸必須大於 0")
    if crop_w_cm <= 0 or crop_h_cm <= 0:
        raise ValueError("裁切尺寸必須大於 0")
    if crop_x < 0 or crop_y < 0:
        raise ValueError("裁切起點不可小於 0")
    if white_cm < 0 or black_cm < 0:
        raise ValueError("邊寬不能為負數")

    files = collect_images(input_path, recursive=recursive)
    files = [src for src in files if not _under(src, out)]
    if not files:
        raise ValueError("未找到可處理的圖片。")

    log(f"輸入: {input_path}")
    log(f"輸出: {out}")
    log(
        f"共 {len(files)} 張 | 擴展 {target_w_cm:g}x{target_h_cm:g} cm | "
        f"裁切 {crop_w_cm:g}x{crop_h_cm:g} cm @ ({crop_x},{crop_y}) | "
        f"白邊 {white_cm} cm / 黑邊 {black_cm} cm\n"
    )

    ok = 0
    total = len(files)
    for i, src in enumerate(files, start=1):
        dest = build_dest(src, input_path, out)
        try:
            process_one(
                src,
                dest,
                fallback_dpi=dpi,
                target_w_cm=target_w_cm,
                target_h_cm=target_h_cm,
                crop_x=crop_x,
                crop_y=crop_y,
                crop_w_cm=crop_w_cm,
                crop_h_cm=crop_h_cm,
                white_cm=white_cm,
                black_cm=black_cm,
                use_image_dpi=use_image_dpi,
                log=log,
            )
            ok += 1
        except Exception as exc:
            log(f"  失敗 {src.name}: {exc}")
        if on_progress is not None:
            on_progress(i, total)

    log(f"\n完成: {ok}/{total}")
    return ok, total


def run_sku_batch(
    parent_dir: Path,
    excel_path: Path,
    output_dir: Path | None = None,
    *,
    dpi: float = DEFAULT_DPI,
    target_w_cm: float = DEFAULT_TARGET_W_CM,
    target_h_cm: float = DEFAULT_TARGET_H_CM,
    crop_x: int = 0,
    crop_y: int = 0,
    white_cm: float = DEFAULT_WHITE_CM,
    black_cm: float = DEFAULT_BLACK_CM,
    use_image_dpi: bool = True,
    log=print,
    on_progress=None,
) -> tuple[int, int]:
    """
    Process one level of SKU-named subfolders under parent_dir.

    Crop size comes from Excel (平台SKU ID + 生产属性尺码). Folders without a
    matching size are skipped and logged.
    """
    parent_dir = parent_dir.resolve()
    if not parent_dir.is_dir():
        raise ValueError(f"輸入必須是父資料夾: {parent_dir}")

    if dpi <= 0:
        raise ValueError("DPI 必須大於 0")
    if target_w_cm <= 0 or target_h_cm <= 0:
        raise ValueError("擴展尺寸必須大於 0")
    if crop_x < 0 or crop_y < 0:
        raise ValueError("裁切起點不可小於 0")
    if white_cm < 0 or black_cm < 0:
        raise ValueError("邊寬不能為負數")

    sku_catalog = load_sku_catalog(excel_path)
    sku_sizes = sku_catalog.sizes
    out = (output_dir or default_output_dir(parent_dir)).resolve()

    sku_dirs = sorted(p for p in parent_dir.iterdir() if p.is_dir())
    if not sku_dirs:
        raise ValueError(f"父資料夾下沒有子資料夾: {parent_dir}")

    jobs: list[tuple[Path, Path, float, float]] = []
    skipped_folders = 0

    for sku_dir in sku_dirs:
        folder_name = sku_dir.name.strip()
        sku = extract_sku_from_folder_name(folder_name)
        if sku is None:
            log(f"略過資料夾 {folder_name}: 無法從名稱辨識 SKU")
            skipped_folders += 1
            continue

        size = sku_sizes.get(sku)
        if size is None:
            if sku in sku_catalog.all_skus:
                reason = "Excel 有此 SKU 但無尺碼"
            else:
                reason = "Excel 中無此 SKU"
            log(f"略過資料夾 {folder_name}（SKU={sku}）: {reason}")
            skipped_folders += 1
            continue

        crop_w_cm, crop_h_cm = size
        files = collect_images(sku_dir, recursive=True)
        files = [src for src in files if not _under(src, out)]
        if not files:
            log(f"略過資料夾 {folder_name}（SKU={sku}）: 無圖片")
            skipped_folders += 1
            continue

        log(
            f"匹配 {folder_name} → SKU {sku} → 裁切 {crop_w_cm:g}x{crop_h_cm:g} cm"
            f"（{len(files)} 張）"
        )
        for src in files:
            # 輸出保留原始資料夾名與相對路徑
            dest = build_dest(src, sku_dir, out / folder_name)
            jobs.append((src, dest, crop_w_cm, crop_h_cm))

    log(f"輸入父目錄: {parent_dir}")
    log(f"Excel: {excel_path.resolve()}")
    log(f"輸出: {out}")
    log(
        f"SKU 對照表 {len(sku_sizes)} 筆 | 子資料夾 {len(sku_dirs)} 個 | "
        f"待處理圖片 {len(jobs)} 張 | 略過資料夾 {skipped_folders} 個"
    )
    log(
        f"擴展 {target_w_cm:g}x{target_h_cm:g} cm | 起點 ({crop_x},{crop_y}) | "
        f"白邊 {white_cm} cm / 黑邊 {black_cm} cm\n"
    )

    if not jobs:
        raise ValueError("沒有可處理的圖片（請確認 SKU 資料夾名稱與 Excel 尺碼）。")

    ok = 0
    total = len(jobs)
    for i, (src, dest, crop_w_cm, crop_h_cm) in enumerate(jobs, start=1):
        # 擴展畫布至少要蓋住裁切尺寸（例如 150x50）
        job_target_w = max(target_w_cm, crop_w_cm)
        job_target_h = max(target_h_cm, crop_h_cm)
        try:
            process_one(
                src,
                dest,
                fallback_dpi=dpi,
                target_w_cm=job_target_w,
                target_h_cm=job_target_h,
                crop_x=crop_x,
                crop_y=crop_y,
                crop_w_cm=crop_w_cm,
                crop_h_cm=crop_h_cm,
                white_cm=white_cm,
                black_cm=black_cm,
                use_image_dpi=use_image_dpi,
                log=log,
            )
            ok += 1
        except Exception as exc:
            log(f"  失敗 {src.parent.name}/{src.name}: {exc}")
        if on_progress is not None:
            on_progress(i, total)

    log(f"\n完成: {ok}/{total}（略過資料夾 {skipped_folders}）")
    return ok, total


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="平鋪擴展 → 裁切 → 白邊/黑邊 → 輸出 TIFF"
    )
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=None,
        help="輸入檔案或資料夾（SKU 批次時為父資料夾）",
    )
    parser.add_argument(
        "--sku-batch",
        type=Path,
        default=None,
        metavar="DIR",
        help="SKU 批次：父資料夾（子資料夾以 SKU 命名）",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="SKU 與尺碼對照 Excel（--sku-batch 時必填）",
    )
    parser.add_argument("-o", "--output", type=Path, default=None, help="輸出目錄")
    parser.add_argument("--dpi", type=float, default=DEFAULT_DPI, help="預設 DPI")
    parser.add_argument("--target-width-cm", type=float, default=DEFAULT_TARGET_W_CM)
    parser.add_argument("--target-height-cm", type=float, default=DEFAULT_TARGET_H_CM)
    parser.add_argument("--crop-x", type=int, default=0)
    parser.add_argument("--crop-y", type=int, default=0)
    parser.add_argument("--crop-width-cm", type=float, default=None)
    parser.add_argument("--crop-height-cm", type=float, default=None)
    parser.add_argument("--white-cm", type=float, default=DEFAULT_WHITE_CM)
    parser.add_argument("--black-cm", type=float, default=DEFAULT_BLACK_CM)
    parser.add_argument("-r", "--recursive", action="store_true")
    parser.add_argument(
        "--force-dpi",
        action="store_true",
        help="強制使用 --dpi，忽略圖片內嵌 DPI",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

    args = parse_args(argv)

    if args.sku_batch is not None:
        if args.excel is None:
            print("錯誤: --sku-batch 需要同時指定 --excel", file=sys.stderr)
            return 1
        try:
            ok, total = run_sku_batch(
                args.sku_batch,
                args.excel,
                args.output,
                dpi=args.dpi,
                target_w_cm=args.target_width_cm,
                target_h_cm=args.target_height_cm,
                crop_x=args.crop_x,
                crop_y=args.crop_y,
                white_cm=args.white_cm,
                black_cm=args.black_cm,
                use_image_dpi=not args.force_dpi,
            )
        except (ValueError, ImportError) as exc:
            print(f"錯誤: {exc}", file=sys.stderr)
            return 1
        return 0 if ok else 1

    if args.input is None:
        print("錯誤: 請指定 input，或使用 --sku-batch 與 --excel", file=sys.stderr)
        return 1

    if args.crop_width_cm is None or args.crop_height_cm is None:
        print("錯誤: 一般模式需要 --crop-width-cm 與 --crop-height-cm", file=sys.stderr)
        return 1

    try:
        ok, total = run_batch(
            args.input,
            args.output,
            dpi=args.dpi,
            target_w_cm=args.target_width_cm,
            target_h_cm=args.target_height_cm,
            crop_x=args.crop_x,
            crop_y=args.crop_y,
            crop_w_cm=args.crop_width_cm,
            crop_h_cm=args.crop_height_cm,
            white_cm=args.white_cm,
            black_cm=args.black_cm,
            recursive=args.recursive,
            use_image_dpi=not args.force_dpi,
        )
    except ValueError as exc:
        print(f"錯誤: {exc}", file=sys.stderr)
        return 1
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
